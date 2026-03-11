"""
Handles MySQL connection, seeding from Excel, and querying cancer mortality data.

Local setup:
    1. Create a .env file with your DB credentials (see .env.example)
    2. Run: python3 database.py  (seeds the database from Excel)

AWS Lambda:
    - Set environment variables in Lambda console
    - seed_database() is never called on Lambda, only query functions are used
"""

import os
import pymysql
from dotenv import load_dotenv

load_dotenv()  # Reads .env locally, ignored on Lambda (uses env vars instead)

# ── Connection config ─────────────────────────────────────────────────────────

DB_CONFIG = {
    "host":     os.getenv("DB_HOST", "localhost"),
    "port":     int(os.getenv("DB_PORT", 3306)),
    "database": os.getenv("DB_NAME", "health_dashboard"),
    "user":     os.getenv("DB_USER", "root"),
    "password": os.getenv("DB_PASSWORD", ""),
    "ssl":      {"ssl": {}},
    "cursorclass": pymysql.cursors.DictCursor
}

# ── Seeding config (local only) ───────────────────────────────────────────────

EXCEL_PATH = os.path.join(
    os.path.dirname(__file__),
    "aihw-can-122-CDiA-2023-Book-2a-Cancer-mortality-and-age-standardised-rates-by-age-5-year-groups.xlsx",
)
SHEET_NAME = "Table S2a.1"
DATA_START = 7

CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS cancer_mortality (
    id                  INT AUTO_INCREMENT PRIMARY KEY,
    data_type           TEXT,
    cancer_group_site   VARCHAR(255),
    year                INT,
    sex                 VARCHAR(20),
    age_group           VARCHAR(30),
    count               INT,
    age_specific_rate   FLOAT,
    asr_2001_aus_std    FLOAT,
    asr_2023_aus        FLOAT,
    asr_who             FLOAT,
    asr_segi            FLOAT,
    icd10_codes         VARCHAR(100)
)
"""

INSERT_SQL = """
INSERT INTO cancer_mortality
    (data_type, cancer_group_site, year, sex, age_group,
     count, age_specific_rate, asr_2001_aus_std,
     asr_2023_aus, asr_who, asr_segi, icd10_codes)
VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
"""

BATCH_SIZE = 5000

# ── Helpers ───────────────────────────────────────────────────────────────────

def get_connection():
    return pymysql.connect(**DB_CONFIG)

def _clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in ("", ". .", ".."):
            return None
        return stripped
    return value

def _to_float(value):
    cleaned = _clean(value)
    if cleaned is None:
        return None
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return None

def _to_int(value):
    cleaned = _clean(value)
    if cleaned is None:
        return None
    try:
        return int(float(cleaned))
    except (ValueError, TypeError):
        return None

# ── Query functions (used by Lambda) ─────────────────────────────────────────

def get_all_skin_mortality_data():
    """
    Returns melanoma death counts grouped by year.
    Used to plot the mortality trend chart on the frontend.
    """
    conn = get_connection()
    with conn.cursor() as cur:
        cur.execute("""
            SELECT year, SUM(count) AS deaths
            FROM cancer_mortality
            WHERE cancer_group_site = 'Melanoma of the skin'
            AND sex = 'Persons'
            GROUP BY year
            ORDER BY year
        """)
        result = cur.fetchall()
    conn.close()
    return result

# ── Seeding (local only, never runs on Lambda) ────────────────────────────────

def seed_database():
    """
    Drops the existing table, recreates it, and seeds from the Excel file.
    Run this once locally to populate RDS before deploying to Lambda.
    """
    print("Connecting to database...")
    conn = get_connection()
    print("Connection successful.")

    with conn.cursor() as cur:
        print("Recreating 'cancer_mortality' table...")
        cur.execute("DROP TABLE IF EXISTS cancer_mortality")
        cur.execute(CREATE_TABLE_SQL)
        print("Table created successfully.")

        print(f"Loading workbook '{EXCEL_PATH}'...")
        from openpyxl import load_workbook
        workbook = load_workbook(filename=EXCEL_PATH, read_only=True)
        sheet = workbook[SHEET_NAME]
        print(f"Workbook loaded. Reading from sheet '{SHEET_NAME}'.")

        data_to_insert = []
        for row in sheet.iter_rows(min_row=DATA_START, values_only=True):
            if not _clean(row[0]):
                continue
            mapped_row = (
                _clean(row[0]),
                _clean(row[1]),
                _to_int(row[2]),
                _clean(row[3]),
                _clean(row[4]),
                _to_int(row[5]),
                _to_float(row[6]),
                _to_float(row[7]),
                _to_float(row[8]),
                _to_float(row[9]),
                _to_float(row[10]),
                _clean(row[11]),
            )
            data_to_insert.append(mapped_row)

            if len(data_to_insert) == BATCH_SIZE:
                cur.executemany(INSERT_SQL, data_to_insert)
                print(f"Inserted {len(data_to_insert)} rows...")
                data_to_insert = []

        if data_to_insert:
            cur.executemany(INSERT_SQL, data_to_insert)
            print(f"Inserted final {len(data_to_insert)} rows.")

    conn.commit()
    conn.close()
    print("Database seeding complete and connection closed.")

if __name__ == "__main__":
    seed_database()

## Also Create a .env.example for GitHub

#Create a file called `.env.example` (safe to commit, no real credentials) so teammates know exactly what to fill in:
#DB_HOST=your-rds-endpoint.rds.amazonaws.com
#DB_PORT=3306
#DB_NAME=health_dashboard
#DB_USER=admin
#DB_PASSWORD=your_password_here
#And make sure `.gitignore` has:
#.env
